"""
Given a sodaJSONObject dataset-structure key, create a skeleton of the dataset structure on the user's filesystem. 
Then pass the path to the skeleton to the validator.
Works within Organize Datasets to allow a user to validate their dataset before uploading it to Pennsieve/Generating it locally.
"""

# create a dummy file in the current directory  
# this is used to check if the script is running in the correct directory

from flask import Flask, abort
import os
import shutil
from xml.dom import InvalidStateErr
import copy
from os.path import expanduser
from pathlib import Path
from sparcur.paths import Path as SparCurPath
from sparcur.simple.validate import main as validate
from sparcur.simple.clean_metadata_files import main as clean_metadata_files
import pandas as pd 
import json 
from namespaces import NamespaceEnum, get_namespace_logger
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import numpy as np
from string import ascii_uppercase
import itertools
import time




app = Flask(__name__)



path = os.path.join(expanduser("~"), "SODA", "skeleton")
completed_jobs_dir = os.path.join(expanduser("~"), "SODA", "completed_jobs")

namespace_logger = get_namespace_logger(NamespaceEnum.VALIDATE_DATASET)

TEMPLATE_PATH = expanduser("~") + "/file_templates"


def has_required_metadata_files(metadata_files_json):
    """
    Checks that the dataset has the required metadata files. These are: dataset_description.xlsx, 
    subjects.xlsx, samples.xlsx, and submission.xlsx. 
    """

    REQUIRED_METADATA_FILES = {
        "submission": False, 
        "dataset_description": False,
        "subjects": False
    }

    for file_name, _ in metadata_files_json.items():
        # strip extension from file_name 
        file_name = file_name.split(".")[0]
        if file_name in REQUIRED_METADATA_FILES:
            REQUIRED_METADATA_FILES[file_name] = True
            

    # return True if all the required metadata files are present
    return all(REQUIRED_METADATA_FILES.values())

def create_validation_error_message(base_message, ds_path):
    error_message = base_message
    if not has_required_metadata_files(ds_path):
        error_message += "Please make sure that you have the required metadata files in your dataset."
    error_message += f"To view the raw report, please see the validation.json file in your SODA folder at {expanduser('~')}/SODA/validation.json"
    return error_message

def create_skeleton(dataset_structure, path):
    """
    Create a skeleton of the dataset structure on the user's filesystem.
    """
    for folder in dataset_structure["folders"]:
        dp = (os.path.join(path, folder)) 
        if not os.path.exists(dp):
            os.mkdir(dp)

        create_skeleton(dataset_structure["folders"][folder], os.path.join(path, folder))
    if "files" in dataset_structure:
        for file_key in dataset_structure["files"]:
            # TODO: If the type is bf then create a generic file with the name of the file key ( and write information to it )
            # if dataset_structure["files"][file_key]["type"] in ["bf"]:
            #     continue
                
            with open(os.path.join(path, file_key), "w") as f:
                f.write("SODA")

def validate_validation_result(export):
    """
        Verifies the integriy of an export retrieved from remote or generated locally.
        Input: export - A dictionary with sparcur.simple.validate or remote validation results.
    """

    # 1. check if the export was not available for retrieval yet even afer waiting for the current maximum wait time
    if export is None:
        raise InvalidStateErr("We had trouble validating your dataset. Please try again. If the problem persists, please contact us at help@fairdataihub.org.")

    # 2. check if the export was a failed validation run TODO: discern between a failed validation run and a dataset with no metadata files 
    inputs = export.get('inputs')

    # NOTE: May not be possible to be None but just in case
    if inputs is None:
        InvalidStateErr("Please add metadata files to your dataset to receive a validation report.")

# # return the errors from the error_path_report that should be shown to the user.
# # as per Tom (developer of the Validator) for any paths (the keys in the Path_Error_Report object)
# # return the ones that do not have any errors in their subpaths. 
# # e.g., If given #/meta and #/meta/technique keys only return #/meta/technique (as this group doesn't have any subpaths)
def parse(error_path_report):

  user_errors = copy.deepcopy(error_path_report)

  keys = error_path_report.keys()

  # go through all paths and store the paths with the longest subpaths for each base 
  # also store matching subpath lengths together
  for k in keys:
    prefix = get_path_prefix(k)

    # check if the current path has inputs as a substring
    if prefix.find("inputs") != -1:
      # as per Tom ignore inputs paths' so
      # remove the given prefix with 'inputs' in its path
      del user_errors[k]
      continue 

    # check for a suffix indicator in the prefix (aka a forward slash at the end of the prefix)
    if prefix[-1] == "/":
      # if so remove the suffix and check if the resulting prefix is an existing path key
      # indicating it can be removed from the errors_for_users dictionary as the current path
      # will be an error in its subpath -- as stated in the function comment we avoid these errors 
      prefix_no_suffix_indicator = prefix[0 : len(prefix) - 1]

      if prefix_no_suffix_indicator in user_errors:
        del user_errors[prefix_no_suffix_indicator]


  
  return user_errors
  

def get_path_prefix(path):
  if path.count('/') == 1:
    # get the entire path as the "prefix" and return it
    return path
  # get the path up to the final "/" and return it as the prefix
  final_slash_idx = path.rfind("/")
  return path[:final_slash_idx + 1]

### Free Form Mode Skeleton Dataset Creation ###

def create_manifests(manifest_struct, path):
   # go through the high level keys
   for key in manifest_struct:
      manifest_obj = manifest_struct[key]

      manifest_df = pd.DataFrame(json.loads(manifest_obj))

      # write to the ~/SODA/skeleton/key folder
      manifest_df.to_excel(f"{path}/{key}/manifest.xlsx", index=False)

def create_metadata_files(metadata_struct, path):
   for metadata_file_name in metadata_struct:
      metadata_obj = metadata_struct[metadata_file_name]

      if metadata_file_name in ["README.txt", "CHANGES.txt"]:
         # write the data to a txt file 
         # TODO: Cannot overwrite existing files that are important 
         with open(f"{path}/{metadata_file_name}", "x") as metadata_file:
            metadata_file.write(metadata_obj)
      else:
        metadata_df = pd.DataFrame(json.loads(metadata_obj))

        metadata_df.to_excel(f"{path}/{metadata_file_name}", index=False, engine="openpyxl")

def create(dataset_structure, manifests_struct, metadata_files, clientUUID):
    """
    Creates a skeleton dataset ( a set of empty data files but with valid metadata files ) of the given soda_json_structure on the local machine.
    Used for validating a user's dataset before uploading it to Pennsieve.
    NOTE: This function is only used for validating datasets ( both local and on Pennsieve ) that are being organized in the Organize Datasets feature of SODA.
    The reason for this being that those datasets may exist in multiple locations on a user's filesystem ( or even on multiple machines ) and therefore cannot be validated 
    until they have been put together in a single location.

    """
    path = os.path.join(expanduser("~"), "SODA", "skeleton")

    # check if the skeleton directory exists
    if not os.path.exists(path):
        # create the skeleton directory
        os.makedirs(path)

    # check if the unique path for the client exists
    path = os.path.join(path, clientUUID)
    if os.path.exists(path):
        # remove the directory and all its contents
        # TODO: ensure no critical/root directories can be deleted. Likely run as non-super user should fix this. 
        shutil.rmtree(path)
    
    # create the directory for the client
    os.mkdir(path)

    create_skeleton(dataset_structure, path)

    # create metadata files 
    namespace_logger.info("{clientUUID}: 2. Creating Metadata Files ( Guided: False )")
    create_metadata_files(metadata_files, path)

    # use pandas to parse the manifest files as data frames then write them to the correct folder
    namespace_logger.info("{clientUUID}: 3. Creating Manifest Files ( Guided: False )")
    create_manifests(manifests_struct, path)

    # clean the metadata files to prevent vaidator hanging on large datasets
    clean_metadata_files(path=SparCurPath(path), cleaned_output_path=SparCurPath(path))

    return path



### Guided Mode Skeleton Dataset Creation ###

subjectsTemplateHeaderList = [
    "subject id",
    "pool id",
    "subject experimental group",
    "age",
    "sex",
    "species",
    "strain",
    "rrid for strain",
    "age category",
    "also in dataset",
    "member of",
    "laboratory internal id",
    "date of birth",
    "age range (min)",
    "age range (max)",
    "body mass",
    "genotype",
    "phenotype",
    "handedness",
    "reference atlas",
    "experimental log file path",
    "experiment date",
    "disease or disorder",
    "intervention",
    "disease model",
    "protocol title",
    "protocol url or doi",
]
samplesTemplateHeaderList = [
    "sample id",
    "subject id",
    "was derived from",
    "pool id",
    "sample experimental group",
    "sample type",
    "sample anatomical location",
    "also in dataset",
    "member of",
    "laboratory internal id",
    "date of derivation",
    "experimental log file path",
    "reference atlas",
    "pathology",
    "laterality",
    "cell type",
    "plane of section",
    "protocol title",
    "protocol url or doi",
]

# transpose a matrix (array of arrays)
def transposeMatrix(matrix):
    return [[matrix[j][i] for j in range(len(matrix))] for i in range(len(matrix[0]))]


# helper function to process custom fields (users add and name them) for subjects and samples files
def processMetadataCustomFields(matrix):
    return [column for column in matrix if any(column[1:])]

# needed to sort subjects and samples table data to match the UI fields
def sortedSubjectsTableData(matrix, fields):
    sortedMatrix = []
    for field in fields:
        for column in matrix:
            if column[0].lower() == field:
                sortedMatrix.append(column)
                break

    customHeaderMatrix = [
        column for column in matrix if column[0].lower() not in fields
    ]

    return (
        np.concatenate((sortedMatrix, customHeaderMatrix)).tolist()
        if customHeaderMatrix
        else sortedMatrix
    )


def excel_columns(start_index=0):
    """
    NOTE: does not support more than 699 contributors/links
    """
    single_letter = list(ascii_uppercase[start_index:])
    two_letter = [a + b for a, b in itertools.product(ascii_uppercase, ascii_uppercase)]
    return single_letter + two_letter


def save_subjects_file(filepath, datastructure):

    source = os.path.join(TEMPLATE_PATH, "subjects.xlsx")

    destination = filepath

    shutil.copyfile(source, destination)
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    transposeDatastructure = transposeMatrix(datastructure)

    mandatoryFields = transposeDatastructure[:11]
    optionalFields = transposeDatastructure[11:]
    refinedOptionalFields = processMetadataCustomFields(optionalFields)

    templateHeaderList = subjectsTemplateHeaderList
    sortMatrix = sortedSubjectsTableData(mandatoryFields, templateHeaderList)

    if refinedOptionalFields:
        refinedDatastructure = transposeMatrix(
            np.concatenate((sortMatrix, refinedOptionalFields))
        )
    else:
        refinedDatastructure = transposeMatrix(sortMatrix)
    
    # 1. delete rows using delete_rows(index, amount=2) -- description and example rows
    # ws1.delete_rows(2, 2)
    # delete all optional columns first (from the template)
    ws1.delete_cols(12, 18)

    # 2. see if the length of datastructure[0] == length of datastructure. If yes, go ahead. If no, add new columns from headers[n-1] onward.
    headers_no = len(refinedDatastructure[0])
    orangeFill = PatternFill(
        start_color="FFD965", end_color="FFD965", fill_type="solid"
    )

    # gevent.sleep(0)
    for column, header in zip(
        excel_columns(start_index=11), refinedDatastructure[0][11:headers_no]
    ):
        cell = column + str(1)
        ws1[cell] = header
        ws1[cell].fill = orangeFill
        ws1[cell].font = Font(bold=True, size=12, name="Calibri")

    # gevent.sleep(0)
    # 3. populate matrices
    for i, item in enumerate(refinedDatastructure):
        if i == 0:
            continue
        for column, j in zip(excel_columns(start_index=0), range(len(item))):
            # import pdb; pdb.set_trace()
            cell = column + str(i + 1)
            ws1[cell] = refinedDatastructure[i][j] or ""
            ws1[cell].font = Font(bold=False, size=11, name="Arial")

    wb.save(destination)


def save_samples_file(filepath, datastructure):
    source = os.pah.join(TEMPLATE_PATH, "samples.xlsx")

    destination = filepath

    shutil.copyfile(source, destination)

    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    transposeDatastructure = transposeMatrix(datastructure)

    mandatoryFields = transposeDatastructure[:9]
    optionalFields = transposeDatastructure[9:]
    refinedOptionalFields = processMetadataCustomFields(optionalFields)

    templateHeaderList = samplesTemplateHeaderList
    sortMatrix = sortedSubjectsTableData(mandatoryFields, templateHeaderList)

    if refinedOptionalFields:
        refinedDatastructure = transposeMatrix(
            np.concatenate((sortMatrix, refinedOptionalFields))
        )
    else:
        refinedDatastructure = transposeMatrix(sortMatrix)

    ws1.delete_cols(10, 15)

    # 2. see if the length of datastructure[0] == length of datastructure. If yes, go ahead. If no, add new columns from headers[n-1] onward.
    headers_no = len(refinedDatastructure[0])
    orangeFill = PatternFill(
        start_color="FFD965", end_color="FFD965", fill_type="solid"
    )
    # gevent.sleep(0)
    for column, header in zip(
        excel_columns(start_index=9), refinedDatastructure[0][9:headers_no]
    ):
        cell = column + str(1)
        ws1[cell] = header
        ws1[cell].fill = orangeFill
        ws1[cell].font = Font(bold=True, size=12, name="Calibri")

    # gevent.sleep(0)
    # 3. populate matrices
    for i, item in enumerate(refinedDatastructure):
        if i == 0:
            continue
        for column, j in zip(excel_columns(start_index=0), range(len(item))):
            # import pdb; pdb.set_trace()
            cell = column + str(i + 1)
            ws1[cell] = refinedDatastructure[i][j] or ""
            ws1[cell].font = Font(bold=False, size=11, name="Arial")

    wb.save(destination)


def rename_headers(workbook, max_len, start_index):
    """
    Rename header columns if values exceed 3. Change Additional Values to Value 4, 5,...
    """

    columns_list = excel_columns(start_index=start_index)
    if max_len >= start_index:
        workbook[columns_list[0] + "1"] = "Value"
        for i, column in zip(range(2, max_len + 1), columns_list[1:]):

            workbook[column + "1"] = f"Value {str(i)}"
            cell = workbook[column + "1"]

            blueFill = PatternFill(
                start_color="9CC2E5", end_color="9CC2E5", fill_type="solid"
            )

            font = Font(bold=True)
            cell.fill = blueFill
            cell.font = font

    else:
        delete_range = len(columns_list) - max_len
        workbook.delete_cols(4 + max_len, delete_range)


def save_submission_file(filepath, val_arr):

    font_submission = Font(name="Calibri", size=14, bold=False)

    source = os.path.join(TEMPLATE_PATH, "submission.xlsx")

    destination = filepath

    try:
        shutil.copyfile(source, destination)
    except FileNotFoundError as e:
        raise e

    # write to excel file
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]
    for column, arr in zip(excel_columns(start_index=2), val_arr):
        ws1[column + "2"] = arr["award"]
        ws1[column + "3"] = arr["milestone"]
        ws1[column + "4"] = arr["date"]

        ws1[column + "2"].font = font_submission
        ws1[column + "3"].font = font_submission
        ws1[column + "4"].font = font_submission

    rename_headers(ws1, len(val_arr), 2)

    wb.save(destination)

    wb.close()

    
def populate_related_info(workbook, val_array):
    ## related links including protocols

    for i, column in zip(range(len(val_array)), excel_columns(start_index=3)):
        workbook[column + "24"] = val_array[i]["description"]
        workbook[column + "25"] = val_array[i]["relation"]
        workbook[column + "26"] = val_array[i]["link"]
        workbook[column + "27"] = val_array[i]["type"]

    return len(val_array)


def populate_contributor_info(workbook, val_array):
    ## award info
    for i, column in zip(
        range(len(val_array["funding"])), excel_columns(start_index=3)
    ):
        workbook[column + "8"] = val_array["funding"][i]

    ### Acknowledgments
    workbook["D9"] = val_array["acknowledgment"]

    ### Contributors
    for contributor, column in zip(
        val_array["contributors"], excel_columns(start_index=3)
    ):
        workbook[column + "19"] = contributor["conName"]
        workbook[column + "20"] = contributor["conID"]
        workbook[column + "21"] = contributor["conAffliation"]
        workbook[column + "22"] = contributor["conRole"]

    return [val_array["funding"], val_array["contributors"]]

def populate_study_info(workbook, val_obj):
    workbook["D11"] = val_obj["study purpose"]
    workbook["D12"] = val_obj["study data collection"]
    workbook["D13"] = val_obj["study primary conclusion"]
    workbook["D17"] = val_obj["study collection title"]

    ## study organ system
    for i, column in zip(
        range(len(val_obj["study organ system"])), excel_columns(start_index=3)
    ):
        workbook[column + "14"] = val_obj["study organ system"][i]
    ## study approach
    for i, column in zip(
        range(len(val_obj["study approach"])), excel_columns(start_index=3)
    ):
        workbook[column + "15"] = val_obj["study approach"][i]
    ## study technique
    for i, column in zip(
        range(len(val_obj["study technique"])), excel_columns(start_index=3)
    ):
        workbook[column + "16"] = val_obj["study technique"][i]

    return max(
        len(val_obj["study organ system"]),
        len(val_obj["study approach"]),
        len(val_obj["study technique"]),
    )


def populate_dataset_info(ws, val_obj):
    ## name, description, type, samples, subjects
    ws["D5"] = val_obj["name"]
    ws["D6"] = val_obj["description"]
    ws["D3"] = val_obj["type"]
    ws["D29"] = val_obj["number of subjects"]
    ws["D30"] = val_obj["number of samples"]

    ## keywords
    for i, column in zip(range(len(val_obj["keywords"])), excel_columns(start_index=3)):
        ws[column + "7"] = val_obj["keywords"][i]

    return val_obj["keywords"]

def fillColor(color, cell):
    colorFill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    cell.fill = colorFill

def grayout_subheaders(workbook, max_len, start_index):
    """
    Gray out sub-header rows for values exceeding 3 (SDS2.0).
    """
    headers_list = ["4", "10", "18", "23", "28"]
    columns_list = excel_columns(start_index=start_index)

    for (i, column), no in itertools.product(zip(range(2, max_len + 1), columns_list[1:]), headers_list):
        cell = workbook[column + no]
        fillColor("B2B2B2", cell)

def grayout_single_value_rows(workbook, max_len, start_index):
    """
    Gray out rows where only single values are allowed. Row number: 2, 3, 5, 6, 9, 11, 12, 13, 17, 29, 30
    """

    columns_list = excel_columns(start_index=start_index)
    row_list = ["2", "3", "5", "6", "9", "11", "12", "13", "17", "29", "30"]
    for (i, column), no in itertools.product(zip(range(2, max_len + 1), columns_list[1:]), row_list):
        cell = workbook[column + no]
        fillColor("CCCCCC", cell)


def save_ds_description_file(
    filepath,
    dataset_str,
    study_str,
    con_str,
    related_info_str,
):
    source = os.path.join(TEMPLATE_PATH, "dataset_description.xlsx")

    destination = filepath

    shutil.copyfile(source, destination)

    # json array to python list
    val_obj_study = study_str
    val_obj_ds = dataset_str
    val_arr_con = con_str
    val_arr_related_info = related_info_str

    # write to excel file
    wb = load_workbook(destination)
    ws1 = wb["Sheet1"]

    ws1["D22"] = ""
    ws1["E22"] = ""
    ws1["D24"] = ""
    ws1["E24"] = ""
    ws1["D25"] = ""
    ws1["E25"] = ""

    keyword_array = populate_dataset_info(ws1, val_obj_ds)

    study_array_len = populate_study_info(ws1, val_obj_study)

    (funding_array, contributor_role_array) = populate_contributor_info(
        ws1, val_arr_con
    )

    related_info_len = populate_related_info(ws1, val_arr_related_info)

    # keywords length
    keyword_len = len(keyword_array)

    # contributors length
    no_contributors = len(contributor_role_array)

    # funding = SPARC award + other funding sources
    funding_len = len(funding_array)

    # obtain length for formatting compliance purpose
    max_len = max(
        keyword_len, funding_len, no_contributors, related_info_len, study_array_len
    )

    rename_headers(ws1, max_len, 3)
    grayout_subheaders(ws1, max_len, 3)
    grayout_single_value_rows(ws1, max_len, 3)

    if ws1["G1"].value == "Value n":
        ws1.delete_cols(7)

    wb.save(destination)



def create_metadata_files_guided(dataset_structure, path, clientUUID):
    # get the table data for subjects and samples 
    subject_table_data = dataset_structure["subjects-table-data"]
    samples_table_data = dataset_structure["samples-table-data"]

    if len(subject_table_data) > 0:
      namespace_logger.info("{clientUUID}: 2.1 Creating subjects.xlsx ( Guided: True )")
      save_subjects_file( path + "/subjects.xlsx", subject_table_data)
    if len(samples_table_data) > 0:
      namespace_logger.info("{clientUUID}: 2.2 Creating samples.xlsx ( Guided: True )")
      save_samples_file( path + "/samples.xlsx", samples_table_data)

    guidedSparcAward = dataset_structure["dataset-metadata"]["shared-metadata"]["sparc-award"];
    guidedMilestones = dataset_structure["dataset-metadata"]["submission-metadata"]["milestones"];
    guidedCompletionDate = dataset_structure["dataset-metadata"]["submission-metadata"]["completion-date"];
    guidedSubmissionMetadataJSON = [{
       "award": guidedSparcAward,
       "date": guidedCompletionDate,
       "milestone": guidedMilestones[0],
    }]
    guidedSubmissionMetadataJSON.extend(
        {
            "award": "",
            "date": "",
            "milestone": milestone,
        }
        for milestone in guidedMilestones
    )
    
    namespace_logger.info("{clientUUID}: 2.3 Creating submission.xlsx ( Guided: True )")
    save_submission_file(path + "/submission.xlsx", guidedSubmissionMetadataJSON)

    # dataset description 
    guidedDatasetInformation = dataset_structure["dataset-metadata"]["description-metadata"]["dataset-information"];

    guidedStudyInformation = dataset_structure["dataset-metadata"]["description-metadata"]["study-information"];

    guidedContributorInformation = dataset_structure["dataset-metadata"]["description-metadata"]["contributor-information"]
    
    if guidedSparcAward not in guidedContributorInformation["funding"]:
      guidedContributorInformation["funding"].insert(0, guidedSparcAward);
  
    contributors = dataset_structure["dataset-metadata"]["description-metadata"]["contributors"];

    guidedContributorInformation["contributors"] = [
      {
        "conAffliation": ', '.join(contributor["conAffliation"]),
        "conID": contributor["conID"],
        "conName": contributor["conName"],
        "conRole": ', '.join(contributor["conRole"]),
        "contributorFirstName": contributor["contributorFirstName"],
        "contributorLastName": contributor["contributorLastName"],
      } for contributor in contributors ]
    

    guidedAdditionalLinks = dataset_structure["dataset-metadata"]["description-metadata"]["additional-links"];
    guidedProtocols = dataset_structure["dataset-metadata"]["description-metadata"]["protocols"];
    allDatasetLinks = guidedAdditionalLinks + guidedProtocols

    namespace_logger.info("{clientUUID}: 2.4 Creating dataset_description.xlsx ( Guided: True )")
    save_ds_description_file(path + "/dataset_description.xlsx", guidedDatasetInformation, guidedStudyInformation, guidedContributorInformation, allDatasetLinks)

    ## README and CHANGES Metadata variables
    guidedReadMeMetadata = dataset_structure["dataset-metadata"]["README"];
    guidedChangesMetadata = dataset_structure["dataset-metadata"]["CHANGES"];

    # create text file called readme.txt in skeleton directory 
    file_path = os.path.join(path, "README.txt")

    namespace_logger.info("{clientUUID}: 2.5 Creating README.txt ( Guided: True )")
    with open(file_path, "w") as f:
        f.write(guidedReadMeMetadata)

    namespace_logger.info("{clientUUID}: 2.6 Creating CHANGES.txt ( Guided: True )")
    file_path = os.path.join(path, "CHANGES.txt")
    if len(guidedChangesMetadata) > 0:
       with open(file_path, "w") as f:
          f.write(guidedChangesMetadata)


def createGuidedMode(soda_json_structure, clientUUID, manifests_struct):
  
  dataset_structure = soda_json_structure["saved-datset-structure-json-obj"]

  path = os.path.join(expanduser("~"), "SODA", "skeleton")

   # check if the skeleton directory exists   
  if not os.path.exists(path):
    # create the skeleton directory
    os.makedirs(path)

  # check if the unique path for the client exists
  path = os.path.join(path, clientUUID)
  if os.path.exists(path):
    # remove the directory and all its contents
    shutil.rmtree(path)
    
  # create the directory for the client
  os.mkdir(path)

  create_skeleton(dataset_structure, path)

  # create metadata files
  namespace_logger.info(f"{clientUUID}: 2. Creating metadata files ( Guided: True ) ")
  create_metadata_files_guided(soda_json_structure, path, clientUUID)

  create_manifests(manifests_struct, path)

  # clean the manifest and metadata files to prevent hanging caused by openpyxl trying to open manifest/metadata files with 
  # excessive amounts of empty rows/columns
  clean_metadata_files(path=SparCurPath(path), cleaned_output_path=SparCurPath(path))




  return path



# validate a local dataset at the target directory 
def val_dataset_local_pipeline(ds_path, clientUUID):
    # convert the path to absolute from user's home directory
    joined_path = os.path.join(expanduser("~"), ds_path.strip())

    # check that the directory exists 
    valid_directory = os.path.isdir(joined_path)

    # give user an error 
    if not valid_directory:
        raise OSError(f"The given directory does not exist: {joined_path}")

    # convert to Path object for Validator to function properly
    norm_ds_path = Path(joined_path)

    # validate the dataset
    blob = None 
    try: 
        blob = validate(norm_ds_path)
    except Exception as e:
       abort(500, e)

    delete_validation_directory(ds_path)

    if 'status' not in blob or 'path_error_report' not in blob['status']:
        namespace_logger.info(f"{clientUUID}: 4.1 Validation Run Incomplete ( Guided: True )")
        return {"parsed_report": {}, "full_report": str(blob), "status": "Incomplete"}
    
    namespace_logger.info(f"{clientUUID}: 4.2 Parsing dataset results( Guided: True ) ")
    
    # peel out the status object 
    status = blob.get('status')

    # peel out the path_error_report object
    path_error_report = status.get('path_error_report')

    # get the errors out of the report that do not have errors in their subpaths (see function comments for the explanation)
    parsed_report = parse(path_error_report)  

    # remove any false positives from the report
    # TODO: Implement the below function
    remove_false_positives(parsed_report, blob)

    return {"parsed_report": parsed_report, "full_report": str(blob), "status": "Complete"}


"""
Delete the validation directory for the given clientUUID.
"""
def delete_validation_directory(clientUUID):
    # check if there is a skeleton dataset directory with this clientUUID as the name
    path = os.path.join(expanduser("~"), "SODA", "skeleton", clientUUID)
    if os.path.exists(path):
        # remove the directory and all its contents
        shutil.rmtree(path)


def remove_false_positives(parsed_report, blob):

    # remove the 'path_metadata' is a required proeprty error message
    if '#/' in parsed_report:
        messages = parsed_report['#/']['messages']
        # remove the message from the messages list with 'path_metadata' as a substring
        for message in messages:
            if 'path_metadata' in message:
                messages.remove(message)
        


    # remove the #/id error from the parsed report if not dealing with Pennsieve
    # TODO: In time we should be able to pull from Pennsieve then validate correctly. Otherwise just pull their ID ourselves and run a regex. 
    if '#/id' in parsed_report:
        del parsed_report['#/id']


if __name__ == '__main__':
    app.run(host='127.0.0.1', port=9000)

    

    

    
      




    